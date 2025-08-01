import os
import sys
import subprocess

# Mostra un messaggio di errore all'utente in GUI se l'installazione fallisce
def mostra_errore_gui(msg):
    try:
        import tkinter as tk
        import tkinter.scrolledtext as ScrolledText
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Errore di installazione", msg)
    except Exception as e:
        print("Errore grave:", msg)

# Tenta di installare un modulo pip
def installa_modulo(modulo, nome_pip=None):
    nome_pip = nome_pip or modulo
    try:
        __import__(modulo)
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", nome_pip])
        except Exception as e:
            mostra_errore_gui(
                f"Non riesco a installare il modulo '{nome_pip}'.\n"
                f"Controlla la connessione a Internet o installa manualmente con:\n\npip install {nome_pip}"
            )
            sys.exit(1)

installa_modulo("openpyxl")
installa_modulo("PIL", nome_pip="pillow")

import tkinter as tk
import tkinter.scrolledtext as ScrolledText
from tkinter import filedialog, messagebox
from tkinter import ttk
from PIL import Image, ImageTk
from datetime import datetime, time
from collections import defaultdict
import openpyxl
from processor import processa_cartella_excel
from diff_cart import confronta_file_cartellini
from count_tot import esegui_count_tot
import threading


os.chdir(os.path.dirname(os.path.abspath(sys.argv[0])))

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Elaborazione Cartellini")
        self.root.geometry("700x700")
        self.root.resizable(False, False)
        self.root.configure(bg="white")

        font_base = ("Segoe UI", 10)
        font_bold = ("Segoe UI", 10, "bold")

        # Tab setup
        self.tab_control = ttk.Notebook(self.root)
        self.tab_elabora = tk.Frame(self.tab_control, bg="white")
        self.tab_count = tk.Frame(self.tab_control, bg="white")
        self.tab_confronta = tk.Frame(self.tab_control, bg="white")
        self.tab_control.add(self.tab_elabora, text="Elabora cartellini")
        self.tab_control.add(self.tab_count, text="Conta Totali")
        self.tab_control.add(self.tab_confronta, text="Confronta")
        self.tab_control.pack(expand=1, fill="both")
        self.setup_tab_count()

        # Logo
        # Carica e mostra il logo cactus in entrambi i tab
        try:
            logo_elabora_img = Image.open("icons/logo.png").resize((80, 80), Image.Resampling.LANCZOS)
            logo_confronta_img = Image.open("icons/logo.png").resize((80, 80), Image.Resampling.LANCZOS)
            self.logo_elabora_img = ImageTk.PhotoImage(logo_elabora_img)
            self.logo_confronta_img = ImageTk.PhotoImage(logo_confronta_img)

            tk.Label(self.tab_elabora, image=self.logo_elabora_img, bg="white").pack(pady=10)
            tk.Label(self.tab_confronta, image=self.logo_confronta_img, bg="white").pack(pady=10)
        except Exception as e:
            print(f"Errore caricamento logo: {e}")


        # Caricamento icone log
        try:
            self.icons = {
                "start": ImageTk.PhotoImage(Image.open("icons/icon_start.png").resize((16, 16), Image.Resampling.LANCZOS)),
                "end": ImageTk.PhotoImage(Image.open("icons/icon_end.png").resize((16, 16), Image.Resampling.LANCZOS)),
                "info": ImageTk.PhotoImage(Image.open("icons/icon_info.png").resize((16, 16), Image.Resampling.LANCZOS)),
                "success": ImageTk.PhotoImage(Image.open("icons/icon_success.png").resize((16, 16), Image.Resampling.LANCZOS)),
                "warning": ImageTk.PhotoImage(Image.open("icons/icon_warning.png").resize((16, 16), Image.Resampling.LANCZOS)),
                "error": ImageTk.PhotoImage(Image.open("icons/icon_error.png").resize((16, 16), Image.Resampling.LANCZOS))
            }
        except Exception as e:
            print(f"Errore nel caricamento delle icone: {e}")
            self.icons = {}

        # CARTELLA ESPORTAZIONI
        frame_cartella = tk.Frame(self.tab_elabora, bg="white")
        frame_cartella.pack(pady=(10, 0), padx=20, anchor="w")
        tk.Label(frame_cartella, text="Cartella con esportazioni:", font=font_base, bg="white").pack(side=tk.LEFT)
        self.btn_cartella = tk.Button(
            frame_cartella, text="Sfoglia...", font=font_bold,
            bg="#e0e0e0", relief=tk.FLAT, cursor="hand2",
            command=self.seleziona_cartella
        )
        self.btn_cartella.pack(side=tk.LEFT, padx=10)

        self.entry_cartella_selezionata = tk.Entry(
            self.tab_elabora, fg="#005bbb", bg="white", font=font_base, state="readonly",
            readonlybackground="white", relief=tk.FLAT, width=90
        )
        self.entry_cartella_selezionata.pack(pady=(0, 10), padx=20, anchor="w")

        # FILE ORE SETTIMANALI
        frame_ore = tk.Frame(self.tab_elabora, bg="white")
        frame_ore.pack(pady=(10, 0), padx=20, anchor="w")
        tk.Label(frame_ore, text="File presenze Factorial:", font=font_base, bg="white").pack(side=tk.LEFT)
        self.btn_ore_settimanali = tk.Button(
            frame_ore, text="Sfoglia...", font=font_bold,
            bg="#e0e0e0", relief=tk.FLAT, cursor="hand2",
            command=self.seleziona_file_ore_settimanali
        )
        self.btn_ore_settimanali.pack(side=tk.LEFT, padx=10)

        self.entry_file_ore_settimanali = tk.Entry(
            self.tab_elabora, fg="#005bbb", bg="white", font=font_base, state="readonly",
            readonlybackground="white", relief=tk.FLAT, width=90
        )
        self.entry_file_ore_settimanali.pack(pady=(0, 10), padx=20, anchor="w")

        # NOME FILE OUTPUT
        frame_output = tk.Frame(self.tab_elabora, bg="white")
        frame_output.pack(padx=20, anchor="w")
        tk.Label(frame_output, text="Nome file output:", font=font_base, bg="white").pack(side=tk.LEFT)
        self.entry_output = tk.Entry(frame_output, width=40, font=font_base, relief=tk.FLAT)
        self.entry_output.insert(0, datetime.now().strftime("Cartellini_%d%B_%H.%M.xlsx"))
        self.entry_output.pack(side=tk.LEFT, padx=10, pady=(0, 10))

        # PROGRESS BAR
        self.progress = ttk.Progressbar(self.tab_elabora, orient="horizontal", mode='determinate')
        self.progress.pack(pady=10, fill=tk.X, padx=20)

        # BOTTONE AVVIA
        self.btn_avvia = tk.Button(
            self.tab_elabora,
            text="Avvia Elaborazione",
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            activebackground="#45a049",
            activeforeground="white",
            relief=tk.FLAT,
            padx=14,
            pady=8,
            cursor="hand2",
            command=self.avvia_elaborazione
        )
        self.btn_avvia.pack(pady=(0, 20))

        # FRAME POST-ELABORAZIONE
        self.frame_post_elaborazione = tk.Frame(self.tab_elabora, bg="white")
        self.frame_post_elaborazione.pack(pady=10)

        # LOG OUTPUT
        self.log_frame = tk.Frame(self.tab_elabora, bg="white")
        self.log_frame.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        self.log = ScrolledText.ScrolledText(
            self.log_frame,
            height=12,
            state='disabled',
            bg="#f5f5f5",
            fg="#333",
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            wrap='word'
        )
        self.log.pack(fill="both", expand=True)

        self.log.tag_config('info', foreground="#555")
        self.log.tag_config('start', foreground="#444")
        self.log.tag_config('success', foreground="#2e7d32", font=("Segoe UI",10,"bold"))
        self.log.tag_config('error', foreground="#d32f2f", font=("Segoe UI",10,"bold"))
        self.log.tag_config('warning', foreground="#ed6c02", font=("Segoe UI",10,"italic"))
        self.log.tag_config('blank', foreground="#333")
        self.log.tag_configure("info", foreground="#222")
        self.log.tag_configure("start", foreground="#444")
        self.log.tag_configure("end", foreground="#444")
        self.log.tag_configure("success", foreground="green", font=font_bold)
        self.log.tag_configure("error", foreground="red", font=font_bold)
        self.log.tag_configure("warning", foreground="orange", font=("Segoe UI", 10, "italic"))
        self.log.tag_configure("blank", foreground="#444")

        # Variabili
        self.cartella = None
        self.file_ore_settimanali = None
        self.file_vecchio = None
        self.file_nuovo = None

        # Setup tab "Confronta"
        self.confronta_due_file_cartellini()


    def seleziona_cartella(self):
        self.cartella = filedialog.askdirectory()
        if self.cartella:
            self.entry_cartella_selezionata.config(state='normal')
            self.entry_cartella_selezionata.delete(0, tk.END)
            self.entry_cartella_selezionata.insert(0, self.cartella)
            self.entry_cartella_selezionata.config(state='readonly')


    def seleziona_file_ore_settimanali(self):
        self.file_ore_settimanali = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if self.file_ore_settimanali:
            self.entry_file_ore_settimanali.config(state='normal')
            self.entry_file_ore_settimanali.delete(0, tk.END)
            self.entry_file_ore_settimanali.insert(0, self.file_ore_settimanali)
            self.entry_file_ore_settimanali.config(state='readonly')


    def logga(self, testo, tag="info"):
        self.log.config(state='normal')
        if tag == "start" and tag in self.icons:
            self.log.image_create(tk.END, image=self.icons[tag])
            self.log.insert(tk.END, " ")
            self.log.insert(tk.END, testo, tag)
            self.log.insert(tk.END, " ")
            self.log.image_create(tk.END, image=self.icons[tag])
            self.log.insert(tk.END, "\n")
        elif tag == "end" and tag in self.icons:
            self.log.insert(tk.END, testo + " ", tag)
            self.log.image_create(tk.END, image=self.icons[tag])
            self.log.insert(tk.END, "\n")
        else:
            if tag in self.icons:
                self.log.image_create(tk.END, image=self.icons[tag])
                self.log.insert(tk.END, " ")
            self.log.insert(tk.END, testo + "\n", tag)
        self.log.see(tk.END)
        self.log.config(state='disabled')

    def logga_elaborazione(self, testo, tag="info"):
        self.log.config(state='normal')
        if tag == "start" and tag in self.icons:
            self.log.image_create(tk.END, image=self.icons[tag])
            self.log.insert(tk.END, " ")
            self.log.insert(tk.END, testo, tag)
            self.log.insert(tk.END, " ")
            self.log.image_create(tk.END, image=self.icons[tag])
            self.log.insert(tk.END, "\n")
        elif tag == "end" and tag in self.icons:
            self.log.insert(tk.END, testo + " ", tag)
            self.log.image_create(tk.END, image=self.icons[tag])
            self.log.insert(tk.END, "\n")
        else:
            if tag in self.icons:
                self.log.image_create(tk.END, image=self.icons[tag])
                self.log.insert(tk.END, " ")
            self.log.insert(tk.END, testo + "\n", tag)
        self.log.see(tk.END)
        self.log.config(state='disabled')

    def logga_confronto(self, testo, tag="info"):
        self.text_output_confronto.config(state='normal')
        if tag == "start" and tag in self.icons:
            self.text_output_confronto.image_create(tk.END, image=self.icons[tag])
            self.text_output_confronto.insert(tk.END, " ")
            self.text_output_confronto.insert(tk.END, testo, tag)
            self.text_output_confronto.insert(tk.END, " ")
            self.text_output_confronto.image_create(tk.END, image=self.icons[tag])
            self.text_output_confronto.insert(tk.END, "\n")
        elif tag == "end" and tag in self.icons:
            self.text_output_confronto.insert(tk.END, testo + " ", tag)
            self.text_output_confronto.image_create(tk.END, image=self.icons[tag])
            self.text_output_confronto.insert(tk.END, "\n")
        else:
            if tag in self.icons:
                self.text_output_confronto.image_create(tk.END, image=self.icons[tag])
                self.text_output_confronto.insert(tk.END, " ")
            self.text_output_confronto.insert(tk.END, testo + "\n", tag)
        self.text_output_confronto.see(tk.END)
        self.text_output_confronto.config(state='disabled')


    def avvia_elaborazione(self):
        if not self.cartella:
            messagebox.showerror("Errore", "Seleziona prima una cartella con i file di esportazione.")
            return

        if not self.file_ore_settimanali:
            messagebox.showerror("Errore", "Seleziona prima il file con le ore settimanali.")
            return

        nome_output = self.entry_output.get().strip()
        if not nome_output:
            messagebox.showerror("Errore", "Inserisci un nome per il file di output.")
            return

        if not nome_output.endswith(".xlsx"):
            nome_output += ".xlsx"

        base, ext = os.path.splitext(nome_output)
        counter = 1
        while os.path.exists(nome_output):
            nome_output = f"{base}_{counter}{ext}"
            counter += 1

        files_excel = [f for f in os.listdir(self.cartella) if f.endswith((".xlsx", ".xlsm"))]
        total_files = len(files_excel)

        if total_files == 0:
            self.logga_elaborazione("Nessun file Excel trovato nella cartella.", "warning")
            return

        self.progress["value"] = 0
        self.progress["maximum"] = total_files
        self.root.update_idletasks()

        self.logga_elaborazione("Perfetto! Dammi un attimo...", "start")

        try:
            def aggiorna_progress(i):
                self.progress["value"] = i
                self.root.update_idletasks()

            processa_cartella_excel(
                self.cartella,
                nome_output,
                self.file_ore_settimanali,
                logger=lambda msg, tag="info": self.logga_elaborazione(msg, tag),
                progress_callback=aggiorna_progress
            )

            self.logga_elaborazione(f"File salvato con successo: {nome_output}", "success")
            self.logga_elaborazione(" ", tag="blank")
            self.logga_elaborazione(" ", tag="blank")
            self.logga_elaborazione("Buon lavoro!", "end")

            # Pulisce eventuali pulsanti precedenti
            for widget in self.frame_post_elaborazione.winfo_children():
                widget.destroy()

            common_style = {
                "bg": "#4CAF50",
                "fg": "white",
                "font": ("Arial", 10, "bold"),
                "padx": 10,
                "pady": 5,
                "cursor": "hand2",
                "activebackground": "#45a049",
                "activeforeground": "white",
                "relief": tk.FLAT
            }

            # Crea i pulsanti base
            btn_apri_file = tk.Button(
                self.frame_post_elaborazione,
                text="Apri file dei cartellini",
                command=lambda: os.startfile(nome_output),
                **common_style
            )
            btn_apri_file.pack(side=tk.LEFT, padx=10)

            btn_verifica = tk.Button(
                self.frame_post_elaborazione,
                text="Verifica inserimento cartellini",
                command=lambda: self.verifica_cartellini(nome_output),
                **common_style
            )
            btn_verifica.pack(side=tk.LEFT, padx=10)

            # Schiarisce il pulsante principale
            self.btn_avvia.config(
                bg="#A5D6A7",
                activebackground="#A5D6A7",
                state="normal",
                relief=tk.FLAT
            )

            # Chiede all'utente se vuole fare il calcolo dei totali
            risposta = messagebox.askyesno(
                "Calcolo Totali",
                "Vuoi eseguire anche il calcolo dei totali sul file appena generato?"
            )

            if risposta:
                try:
                    self.progress["value"] = 0
                    self.progress["maximum"] = 100
                    self.root.update_idletasks()

                    def progress_callback(frazione):
                        self.progress["value"] = int(frazione * 100)
                        self.root.update_idletasks()

                    output_count_tot = nome_output.replace(".xlsx", "_totali.xlsx")
                    base, ext = os.path.splitext(output_count_tot)
                    counter = 1
                    while os.path.exists(output_count_tot):
                        output_count_tot = f"{base}_{counter}{ext}"
                        counter += 1

                    esegui_count_tot(nome_output, output_count_tot, progress_callback)
                    self.logga_elaborazione(f"Totali calcolati e salvati in: {output_count_tot}", "success")

                    # Aggiunge il pulsante "Apri Conteggi"
                    btn_apri_totali = tk.Button(
                        self.frame_post_elaborazione,
                        text="Apri Conteggi",
                        command=lambda: os.startfile(output_count_tot),
                        bg="#FF9800",
                        fg="white",
                        font=("Arial", 10, "bold"),
                        relief=tk.FLAT,
                        activebackground="#FB8C00",
                        activeforeground="white",
                        padx=10,
                        pady=5,
                        cursor="hand2"
                    )
                    btn_apri_totali.pack(side=tk.LEFT, padx=10)

                except Exception as e:
                    self.logga_elaborazione(f"Errore nel calcolo dei totali: {e}", "error")
                    messagebox.showerror("Errore Totali", str(e))

        except Exception as e:
            self.logga_elaborazione(f"Errore: {e}", "error")
            messagebox.showerror("Errore durante l'elaborazione", str(e))
        finally:
            self.progress["value"] = 0

    

    def verifica_cartellini(self, file_output):
        risultati = defaultdict(int)
        try:
            wb = openpyxl.load_workbook(file_output)
            for sheet in wb.worksheets:
                ore_col = None
                for col in sheet.iter_cols(min_row=2, max_row=2):
                    cell = col[0]
                    if cell.value and isinstance(cell.value, str) and "ore lavorate" in cell.value.lower():
                        ore_col = cell.column
                        break
                if ore_col:
                    for row in sheet.iter_rows(min_row=3):
                        data_cell = row[0]
                        if data_cell.value in (None, "", "Totale"):
                            continue
                        cell = row[ore_col - 1]
                        val = cell.value
                        if (
                            val in (None, "", 0, 0.0)
                            or (isinstance(val, str) and val.strip() in ("0", "0:00", "00:00"))
                            or (isinstance(val, time) and val == time(0, 0))
                        ):
                            risultati[sheet.title] += 1

            if risultati:
                output = sorted(risultati.items(), key=lambda x: x[1], reverse=True)
                win = tk.Toplevel(self.root)
                win.title("Verifica cartellini")
                win.geometry("800x450")
                win.resizable(True, True)

                filtro_frame = tk.Frame(win)
                filtro_frame.pack(pady=5)
                tk.Label(filtro_frame, text="Mostra solo dipendenti con più di").pack(side=tk.LEFT)
                soglia_min_var = tk.IntVar(value=0)
                tk.Entry(filtro_frame, textvariable=soglia_min_var, width=5).pack(side=tk.LEFT)
                tk.Label(filtro_frame, text="e meno di").pack(side=tk.LEFT)
                soglia_max_var = tk.IntVar(value=100)
                tk.Entry(filtro_frame, textvariable=soglia_max_var, width=5).pack(side=tk.LEFT)
                tk.Label(filtro_frame, text="giorni mancanti").pack(side=tk.LEFT, padx=(5, 20))

                tk.Label(filtro_frame, text="Filtra per nome:").pack(side=tk.LEFT)
                filtro_nome_var = tk.StringVar()
                tk.Entry(filtro_frame, textvariable=filtro_nome_var, width=15).pack(side=tk.LEFT)

                frame_testo = tk.Frame(win)
                frame_testo.pack(fill=tk.BOTH, expand=True)
                text = tk.Text(frame_testo, wrap=tk.WORD)
                scrollbar = tk.Scrollbar(frame_testo, command=text.yview)
                text.configure(yscrollcommand=scrollbar.set)
                text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

                def aggiorna_output():
                    text.config(state='normal')
                    text.delete(1.0, tk.END)

                    header = "Dipendenti con giorni non compilati:"
                    text.insert(tk.END, header)
                    conta = 0
                    try:
                        min_val = int(soglia_min_var.get())
                    except (tk.TclError, ValueError):
                        min_val = 0
                    try:
                        max_val = int(soglia_max_var.get())
                    except (tk.TclError, ValueError):
                        max_val = 100

                    for nome, count in output:
                        if (
                            count > min_val and count < max_val and
                            filtro_nome_var.get().lower() in nome.lower()
                        ):
                            line = f"{nome}: {count} giorni"
                            text.insert(tk.END, line)
                            conta += 1

                    text.insert(tk.END, f"Totale dipendenti trovati: {conta}")
                    text.config(state='disabled')

                soglia_min_var.trace_add("write", lambda *_: aggiorna_output())
                soglia_max_var.trace_add("write", lambda *_: aggiorna_output())
                filtro_nome_var.trace_add("write", lambda *_: aggiorna_output())

                aggiorna_output()

                btn = tk.Button(win, text="Chiudi", command=win.destroy)
                btn.pack(pady=5)
            else:
                messagebox.showinfo("Verifica cartellini", "Tutti i cartellini risultano compilati.")

        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante la verifica: {e}")

    def seleziona_file(self, tipo):
        path = filedialog.askopenfilename(title=f"Seleziona file cartellini {tipo.upper()}", filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if tipo == "vecchio":
            self.file_vecchio = path
            self.lbl_file_vecchio.config(text=path)
        elif tipo == "nuovo":
            self.file_nuovo = path
            self.lbl_file_nuovo.config(text=path)

    def esegui_confronto(self):
        self.progressbar_confronta["value"] = 0
        self.progressbar_confronta.update()

        file_vecchio = self.file_vecchio
        file_nuovo = self.file_nuovo
        output_path = self.entry_output_confronto.get().strip()

        if not file_vecchio or not os.path.isfile(file_vecchio):
            messagebox.showerror("Errore", "Seleziona un file cartellini VECCHIO valido.")
            return

        if not file_nuovo or not os.path.isfile(file_nuovo):
            messagebox.showerror("Errore", "Seleziona un file cartellini NUOVO valido.")
            return

        if not output_path:
            messagebox.showerror("Errore", "Inserisci un nome per il file di confronto.")
            return

        if not output_path.endswith(".xlsx"):
            output_path += ".xlsx"

        base, ext = os.path.splitext(output_path)
        counter = 1
        while os.path.exists(output_path):
            output_path = f"{base}_{counter}{ext}"
            counter += 1

        try:
            self.logga_confronto("Inizio confronto file cartellini...", "start")
            self.progressbar_confronta["value"] = 50
            self.progressbar_confronta.update()
            confronta_file_cartellini(file_vecchio, file_nuovo, output_path)
            self.logga_confronto(f"✅ Confronto completato. File salvato in: {output_path}", "success")
            self.progressbar_confronta["value"] = 100
            self.progressbar_confronta.update()

            # 🔄 Rimuove vecchi pulsanti o messaggi
            for widget in self.frame_post_confronto.winfo_children():
                widget.destroy()

            # 🟦 Schiarisce visivamente il pulsante "Esegui Confronto"
            self.btn_esegui_confronto.config(
                bg="#90CAF9",  # azzurro chiaro
                activebackground="#90CAF9",
                state="normal",
                relief=tk.FLAT
            )

            # ✅ Aggiunge bottone per aprire il file
            btn_apri_confronto = tk.Button(
                self.frame_post_confronto,
                text="Apri file confronto",
                command=lambda: os.startfile(output_path),
                bg="#2196F3",
                fg="white",
                font=("Arial", 10, "bold"),
                relief=tk.FLAT,
                activebackground="#1976D2",
                activeforeground="white",
                padx=10,
                pady=5,
                cursor="hand2"
            )
            btn_apri_confronto.pack(pady=5)

        except Exception as e:
            self.logga_confronto(f"❌ Errore durante il confronto: {e}", "error")
            messagebox.showerror("Errore", f"Errore durante il confronto: {e}")

    def confronta_due_file_cartellini(self):
        font_base = ("Segoe UI", 10)
        font_bold = ("Segoe UI", 10, "bold")

        # File VECCHIO
        frame_file_vecchio = tk.Frame(self.tab_confronta, bg="white")
        frame_file_vecchio.pack(pady=5, padx=20, anchor="w")
        tk.Label(frame_file_vecchio, text="File cartellini VECCHIO:", font=font_base, bg="white").pack(side=tk.LEFT)
        btn_file_vecchio = tk.Button(
            frame_file_vecchio, text="Sfoglia...", font=font_bold,
            bg="#e0e0e0", relief=tk.FLAT, cursor="hand2",
            command=lambda: self.seleziona_file("vecchio")
        )
        btn_file_vecchio.pack(side=tk.LEFT, padx=10)

        self.lbl_file_vecchio = tk.Label(self.tab_confronta, text="", fg="#005bbb", bg="white", font=font_base)
        self.lbl_file_vecchio.pack(pady=(0, 10), padx=20, anchor="w")

        # File NUOVO
        frame_file_nuovo = tk.Frame(self.tab_confronta, bg="white")
        frame_file_nuovo.pack(pady=5, padx=20, anchor="w")
        tk.Label(frame_file_nuovo, text="File cartellini NUOVO:", font=font_base, bg="white").pack(side=tk.LEFT)
        btn_file_nuovo = tk.Button(
            frame_file_nuovo, text="Sfoglia...", font=font_bold,
            bg="#e0e0e0", relief=tk.FLAT, cursor="hand2",
            command=lambda: self.seleziona_file("nuovo")
        )
        btn_file_nuovo.pack(side=tk.LEFT, padx=10)

        self.lbl_file_nuovo = tk.Label(self.tab_confronta, text="", fg="#005bbb", bg="white", font=font_base)
        self.lbl_file_nuovo.pack(pady=(0, 10), padx=20, anchor="w")

        # Output filename
        frame_output_confronto = tk.Frame(self.tab_confronta, bg="white")
        frame_output_confronto.pack(padx=20, anchor="w")
        tk.Label(frame_output_confronto, text="Nome file di confronto:", font=font_base, bg="white").pack(side=tk.LEFT)
        self.entry_output_confronto = tk.Entry(frame_output_confronto, width=40, font=font_base, relief=tk.FLAT)
        self.entry_output_confronto.pack(side=tk.LEFT, padx=10, pady=(0, 10))

        # Nome di default
        default_name = datetime.now().strftime("Confronto_%d%B_%H.%M.xlsx")
        self.entry_output_confronto.insert(0, default_name)

        # Progress bar
        self.progressbar_confronta = ttk.Progressbar(self.tab_confronta, orient="horizontal", mode='determinate')
        self.progressbar_confronta.pack(pady=10, fill=tk.X, padx=20)

        # Bottone Confronta
        self.btn_esegui_confronto = tk.Button(
            self.tab_confronta,
            text="Esegui Confronto",
            command=self.esegui_confronto,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            activebackground="#2196F3",
            activeforeground="white",
            relief=tk.FLAT,
            padx=14,
            pady=8,
            cursor="hand2"
        )
        self.btn_esegui_confronto.pack(pady=(0, 20))

        # Post confronto
        self.frame_post_confronto = tk.Frame(self.tab_confronta, bg="white")
        self.frame_post_confronto.pack(pady=10)


        # Log box
        self.text_output_confronto = ScrolledText.ScrolledText(
            self.tab_confronta, height=10, bg="#f5f5f5", fg="#333",
            font=("Segoe UI", 10), relief=tk.FLAT, state='disabled', wrap='word'
        )
        self.text_output_confronto.pack(padx=20, pady=(10, 20), fill="both", expand=True)

        self.text_output_confronto.tag_config('info', foreground="#555")
        self.text_output_confronto.tag_config('start', foreground="#444")
        self.text_output_confronto.tag_config('success', foreground="#2e7d32", font=("Segoe UI",10,"bold"))
        self.text_output_confronto.tag_config('error', foreground="#d32f2f", font=("Segoe UI",10,"bold"))
        self.text_output_confronto.tag_config('warning', foreground="#ed6c02", font=("Segoe UI",10,"italic"))
        self.text_output_confronto.tag_config('blank', foreground="#333")


    def setup_tab_count(self):
        font_base = ("Segoe UI", 10)
        font_bold = ("Segoe UI", 10, "bold")

        # Logo
        try:
            logo_img = Image.open("icons/logo.png").resize((80, 80), Image.Resampling.LANCZOS)
            self.logo_count_img = ImageTk.PhotoImage(logo_img)
            tk.Label(self.tab_count, image=self.logo_count_img, bg="white").pack(pady=10)
        except Exception as e:
            print(f"Errore caricamento logo: {e}")

        # Selezione file
        frame_file = tk.Frame(self.tab_count, bg="white")
        frame_file.pack(pady=5, padx=20, anchor="w")

        tk.Label(frame_file, text="File da elaborare:", font=font_base, bg="white").pack(side=tk.LEFT)
        btn_sfoglia = tk.Button(
            frame_file, text="Sfoglia...", font=font_bold,
            bg="#e0e0e0", relief=tk.FLAT, cursor="hand2",
            command=self.seleziona_file_count
        )
        btn_sfoglia.pack(side=tk.LEFT, padx=10)

        self.lbl_file_count = tk.Label(self.tab_count, text="", fg="#005bbb", bg="white", font=font_base)
        self.lbl_file_count.pack(pady=(0, 10), padx=20, anchor="w")

        # Output filename
        frame_output = tk.Frame(self.tab_count, bg="white")
        frame_output.pack(padx=20, anchor="w")
        tk.Label(frame_output, text="Nome file output:", font=font_base, bg="white").pack(side=tk.LEFT)

        self.entry_output_count = tk.Entry(frame_output, width=40, font=font_base, relief=tk.FLAT)
        default_name = datetime.now().strftime("Totali_%d%B_%H.%M.xlsx")
        self.entry_output_count.insert(0, default_name)
        self.entry_output_count.pack(side=tk.LEFT, padx=10, pady=(0, 10))

        # Progress bar
        self.progressbar_count = ttk.Progressbar(self.tab_count, orient="horizontal", mode='determinate')
        self.progressbar_count.pack(pady=10, fill=tk.X, padx=20)

        # Bottone avvia
        self.btn_avvia_count_tot = tk.Button(
            self.tab_count,
            text="Esegui Conta Totali",
            command=self.avvia_count_tot,
            bg="#FB8C00",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            activebackground="#FB8C00",
            activeforeground="white",
            relief=tk.FLAT,
            padx=14,
            pady=8,
            cursor="hand2"
        )
        self.btn_avvia_count_tot.pack(pady=(0, 20))

        # Post-elaborazione
        self.frame_post_count = tk.Frame(self.tab_count, bg="white")
        self.frame_post_count.pack(pady=10)

        # Log
        self.text_output_count = ScrolledText.ScrolledText(
            self.tab_count, height=10, bg="#f5f5f5", fg="#333",
            font=("Segoe UI", 10), relief=tk.FLAT, state='disabled', wrap='word'
        )
        self.text_output_count.pack(padx=20, pady=(10, 20), fill="both", expand=True)

        for tag, color in {
            'info': "#555", 'start': "#444", 'success': "#2e7d32",
            'error': "#d32f2f", 'warning': "#ed6c02", 'blank': "#333"
        }.items():
            self.text_output_count.tag_config(tag, foreground=color)


    def seleziona_file_count(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        self.file_count_path = path
        self.lbl_file_count.config(text=path)


    def avvia_count_tot(self):
        file_input = getattr(self, "file_count_path", None)
        file_output = self.entry_output_count.get().strip()

        if not file_input or not os.path.isfile(file_input):
            messagebox.showerror("Errore", "Seleziona un file Excel valido.")
            return

        if not file_output:
            messagebox.showerror("Errore", "Inserisci un nome per il file di output.")
            return

        if not file_output.endswith(".xlsx"):
            file_output += ".xlsx"

        base, ext = os.path.splitext(file_output)
        counter = 1
        while os.path.exists(file_output):
            file_output = f"{base}_{counter}{ext}"
            counter += 1

        self.progressbar_count["value"] = 0
        self.progressbar_count["maximum"] = 100
        self.root.update_idletasks()

        self.text_output_count.config(state='normal')
        self.text_output_count.delete(1.0, tk.END)
        self.text_output_count.insert(tk.END, "Inizio elaborazione...\n", "start")
        self.root.update_idletasks()

        # 🟦 Schiarisce visivamente il pulsante "Esegui Conta Totali"
        self.btn_avvia_count_tot.config(
            bg="#FFCC80",  # colore attivo (chiaro)
            activebackground="#FFCC80",
            state="normal",
            relief=tk.FLAT
        )
        
        self.root.update_idletasks()

        # Pulisce eventuali pulsanti precedenti
        for widget in self.frame_post_count.winfo_children():
            widget.destroy()

        def run():
            def progress_callback(frazione):
                self.progressbar_count["value"] = int(frazione * 100)
                self.root.update_idletasks()

            try:
                esegui_count_tot(file_input, file_output, progress_callback=progress_callback)
                self.progressbar_count["value"] = 100
                self.root.update_idletasks()

                self.text_output_count.insert(tk.END, f"Operazione completata.\nFile salvato in: {file_output}", "success")

                btn_apri_file = tk.Button(
                    self.frame_post_count,
                    text="Apri file totali",
                    command=lambda: os.startfile(file_output),
                    bg="#FF9800",
                    fg="white",
                    font=("Arial", 10, "bold"),
                    relief=tk.FLAT,
                    activebackground="#FB8C00",
                    activeforeground="white",
                    padx=10,
                    pady=5,
                    cursor="hand2"
                )
                btn_apri_file.pack(pady=5)

            except Exception as e:
                self.text_output_count.insert(tk.END, f"Errore durante l'elaborazione: {e}", "error")
                messagebox.showerror("Errore", str(e))
            finally:
                self.text_output_count.config(state='disabled')


        # Avvia in thread separato per non bloccare la GUI
        threading.Thread(target=run).start()





if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()