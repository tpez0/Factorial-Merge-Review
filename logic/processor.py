import os
import re
import random
from datetime import timedelta, datetime, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy
from openpyxl.utils import get_column_letter

def parse_time(value):
    if isinstance(value, str):
        value = value.strip()
        match = re.match(r"^(\d+):(\d+)$", value)
        if match:
            hours, minutes = map(int, match.groups())
            return timedelta(hours=hours, minutes=minutes)
        # Tenta interpretazione come numero float in ore
        try:
            return timedelta(hours=float(value))
        except ValueError:
            return timedelta(0)
    elif isinstance(value, (int, float)):
        return timedelta(hours=value)
    elif isinstance(value, time):
        return timedelta(hours=value.hour, minutes=value.minute)
    elif isinstance(value, datetime):
        return timedelta(hours=value.hour, minutes=value.minute)
    return timedelta(0)

def format_timedelta(td):
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{hours}:{minutes:02d}"

def normalizza_nome(nome):
    if not nome:
        return ""
    try:
        if isinstance(nome, bytes):
            s_nome = nome.decode('utf-8', errors='ignore')
        else:
            s_nome = str(nome)
        
        normalized_name = re.sub(r"\s+", " ", s_nome.strip().lower().replace("’", "'").replace("`", "'"))
        return normalized_name
    except Exception as e:
        print(f"Errore durante la normalizzazione del nome '{nome}': {e}")
        return str(nome).strip().lower()

def trova_e_somma_celle_rosse_settimanali(ws, col_somma, rosso_chiaro_fill):
    celle_rosse = []
    somma_totale = timedelta(0)
    
    for r in range(5, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_somma)
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == rosso_chiaro_fill.fgColor.rgb:
            celle_rosse.append(cell)
            parsed_time = parse_time(cell.value)
            somma_totale += parsed_time
            
    return somma_totale, celle_rosse

def processa_cartella_excel(cartella_excel, file_output, file_ore_settimanali=None, logger=print, progress_callback=None):
    rosso_chiaro = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    rosso_scuro = PatternFill(start_color="FF6666", end_color="FFB3B3B3", fill_type="solid")
    giallo_chiaro = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
    verde_discrepanza = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")
    
    wb_output = Workbook()
    wb_output.remove(wb_output.active) 

    files_excel = [f for f in os.listdir(cartella_excel) if f.lower().endswith(('.xlsx', '.xlsm'))]
    totale = len(files_excel)
    completati = 0

    intestazioni_fisse = ["Data", "Ore lavorate"]
    intestazioni_opzionali_entrata = ["Orario di entrata"]
    intestazioni_opzionali_uscita = ["Orario d'uscita"]
    
    # MODIFICA QUI: Aggiungi "Tempo tracciato" alle intestazioni totali possibili per la mappatura
    intestazioni_totali_possibili_mappatura = intestazioni_fisse + \
                                               intestazioni_opzionali_entrata + \
                                               intestazioni_opzionali_uscita + \
                                               ["Tempo tracciato"] # AGGIUNTA

    intestazione_somme_settimanali = "Somme ore settimanali" 

    info_totali_per_foglio = {} 

    for nome_file in files_excel:
        cell_ore_sett_value = "" 
        
        percorso_file = os.path.join(cartella_excel, nome_file)

        try:
            wb = load_workbook(percorso_file)
            ws_orig = wb[wb.sheetnames[0]] 

            header_row = 5
            def normalizza(s):
                if s is None:
                    return ""
                return str(s).strip().lower().replace("’", "'").replace("  ", " ")

            col_map_all_instances = {}
            for col_idx, cell in enumerate(ws_orig[header_row], start=1):
                val_norm = normalizza(cell.value)
                original_header_name = None
                for h in intestazioni_totali_possibili_mappatura:
                    if normalizza(h) == val_norm:
                        original_header_name = h
                        break
                
                if original_header_name:
                    col_map_all_instances.setdefault(original_header_name, []).append(col_idx)

            if not all(h in col_map_all_instances for h in intestazioni_fisse):
                logger(f"⚠️ File {nome_file}: intestazioni fisse mancanti, saltato.")
                completati += 1
                if progress_callback:
                    progress_callback(completati)
                continue

            intestazioni_output = []
            col_idx_output_map = {} 

            all_entry_exit_cols = []
            if intestazioni_opzionali_entrata[0] in col_map_all_instances:
                for idx in col_map_all_instances[intestazioni_opzionali_entrata[0]]:
                    all_entry_exit_cols.append(("Entrata", idx))
            if intestazioni_opzionali_uscita[0] in col_map_all_instances:
                for idx in col_map_all_instances[intestazioni_opzionali_uscita[0]]:
                    all_entry_exit_cols.append(("Uscita", idx))
            
            all_entry_exit_cols.sort(key=lambda x: x[1])

            for h_fixed in intestazioni_fisse:
                if h_fixed == "Ore lavorate":
                    continue
                intestazioni_output.append(h_fixed)

            entrata_count = 0
            uscita_count = 0
            for col_type, _ in all_entry_exit_cols:
                if col_type == "Entrata":
                    entrata_count += 1
                    intestazioni_output.append(f"Orario di entrata {entrata_count}")
                elif col_type == "Uscita":
                    uscita_count += 1
                    intestazioni_output.append(f"Orario d'uscita {uscita_count}")
            
            if "Ore lavorate" in intestazioni_fisse and "Ore lavorate" in col_map_all_instances:
                intestazioni_output.append("Ore lavorate")
            
            # AGGIUNTA: Se "Tempo tracciato" esiste nel file originale, aggiungilo alle intestazioni di output
            if "Tempo tracciato" in col_map_all_instances:
                intestazioni_output.append("Tempo tracciato") # AGGIUNTA


            for idx, header in enumerate(intestazioni_output):
                col_idx_output_map[header] = idx + 1

            base_name = os.path.splitext(nome_file)[0]
            if "_full" in base_name:
                nome_foglio = base_name.split("_full")[0].strip()
            else:
                nome_foglio = base_name
            
            nome_foglio = nome_foglio[:31]

            ws_nuovo = wb_output.create_sheet(title=nome_foglio)

            # --- Lettura file ore settimanali ---
            mappa_ore_settimanali = {}
            if file_ore_settimanali:
                try:
                    wb_ore = load_workbook(file_ore_settimanali)
                    ws_ore = wb_ore.active

                    col_fullname = None
                    col_ore = None

                    for idx, cell in enumerate(ws_ore[1], start=1):
                        val = str(cell.value).strip().lower() if cell.value else ""
                        if val == "full_name":
                            col_fullname = idx
                        elif val == "weekly_workschedule_hours":
                            col_ore = idx

                    if col_fullname is None or col_ore is None:
                        logger("⚠️ Colonne 'full_name' o 'weekly_workschedule_hours' mancanti nel file ore settimanali.")
                    else:
                        for row in ws_ore.iter_rows(min_row=2):
                            nome_cell = row[col_fullname - 1].value
                            ore_cell_value = row[col_ore - 1].value 
                            if nome_cell and ore_cell_value is not None:
                                nome_norm = normalizza_nome(nome_cell)
                                
                                parsed_td = parse_time(ore_cell_value) 
                                formatted_ore_str = format_timedelta(parsed_td) 

                                mappa_ore_settimanali[nome_norm] = formatted_ore_str 

                except Exception as e:
                    logger(f"⚠️ Errore nella lettura del file ore settimanali: {e}")
                    import traceback
                    logger(traceback.format_exc())

            for c in [1, 2]: 
                if ws_orig.cell(row=3, column=c).value is not None:
                    cella_orig = ws_orig.cell(row=3, column=c)
                    cella_nuova = ws_nuovo.cell(row=3, column=c, value=cella_orig.value)
                    if cella_orig.has_style:
                        cella_nuova.font = copy(cella_orig.font)
                        cella_nuova.border = copy(cella_orig.border)
                        cella_nuova.fill = copy(cella_orig.fill)
                        cella_nuova.number_format = copy(cella_orig.number_format)
                        cella_nuova.protection = copy(cella_orig.protection)
                        cella_nuova.alignment = copy(cella_orig.alignment)

            current_output_col = 1
            for header_name_output in intestazioni_output:
                cella_nuova_header = ws_nuovo.cell(row=4, column=current_output_col, value=header_name_output)
                
                original_header_base_name = re.sub(r' \d+$', '', header_name_output) 
                
                original_col_idx_for_style = None

                if original_header_base_name == "Orario di entrata" and "Entrata" in [item[0] for item in all_entry_exit_cols]:
                    match = re.search(r' (\d+)$', header_name_output)
                    if match:
                        suffix_num = int(match.group(1))
                        if 1 <= suffix_num <= len([item for item in all_entry_exit_cols if item[0] == "Entrata"]):
                            count_e = 0
                            for etype, eidx in all_entry_exit_cols:
                                if etype == "Entrata":
                                    count_e += 1
                                if count_e == suffix_num:
                                    original_col_idx_for_style = eidx
                                    break
                    if not original_col_idx_for_style and intestazioni_opzionali_entrata[0] in col_map_all_instances: 
                        original_col_idx_for_style = col_map_all_instances[intestazioni_opzionali_entrata[0]][0]

                elif original_header_base_name == "Orario d'uscita" and "Uscita" in [item[0] for item in all_entry_exit_cols]:
                    match = re.search(r' (\d+)$', header_name_output)
                    if match:
                        suffix_num = int(match.group(1))
                        if 1 <= suffix_num <= len([item for item in all_entry_exit_cols if item[0] == "Uscita"]):
                            count_u = 0
                            for utype, uidx in all_entry_exit_cols:
                                if utype == "Uscita":
                                    count_u += 1
                                if count_u == suffix_num:
                                    original_col_idx_for_style = uidx
                                    break
                    if not original_col_idx_for_style and intestazioni_opzionali_uscita[0] in col_map_all_instances: 
                        original_col_idx_for_style = col_map_all_instances[intestazioni_opzionali_uscita[0]][0]
                
                elif original_header_base_name in col_map_all_instances: 
                    original_col_idx_for_style = col_map_all_instances[original_header_base_name][0]
                
                if original_col_idx_for_style:
                    cella_orig_for_style = ws_orig.cell(row=5, column=original_col_idx_for_style)
                    if cella_orig_for_style.has_style:
                        cella_nuova_header.font = copy(cella_orig_for_style.font)
                        cella_nuova_header.border = copy(cella_orig_for_style.border)
                        cella_nuova_header.fill = copy(cella_orig_for_style.fill)
                        cella_nuova_header.number_format = copy(cella_orig_for_style.number_format)
                        original_alignment = copy(cella_orig_for_style.alignment)
                        original_alignment.wrap_text = True
                        cella_nuova_header.alignment = original_alignment 
                        cella_nuova_header.protection = copy(cella_orig_for_style.protection)
                else:
                    cella_nuova_header.font = Font(bold=True)
                    cella_nuova_header.alignment = Alignment(wrap_text=True) 
                
                current_output_col += 1

            nuova_riga = 5
            riga_map = {} 
            ultima_domenica = 0 

            for i, row_cells in enumerate(ws_orig.iter_rows(), start=1):
                if i <= header_row:
                    continue 

                evidenzia_domenica = False # Renamed to avoid conflict with `evidenzia` from user's original logic
                
                # Variabili per i valori delle ore lavorate e tempo tracciato nel ciclo corrente
                ore_lavorate_td_current_row = timedelta(0)
                tempo_tracciato_td_current_row = timedelta(0)

                # Recupera i valori "Ore lavorate" e "Tempo tracciato" dalla riga originale
                col_ore_lavorate_orig_idx = col_map_all_instances["Ore lavorate"][0] if "Ore lavorate" in col_map_all_instances else None
                col_tempo_tracciato_orig_idx = col_map_all_instances["Tempo tracciato"][0] if "Tempo tracciato" in col_map_all_instances else None

                if col_ore_lavorate_orig_idx:
                    ore_lavorate_val_orig = row_cells[col_ore_lavorate_orig_idx - 1].value
                    ore_lavorate_td_current_row = parse_time(ore_lavorate_val_orig)
                
                if col_tempo_tracciato_orig_idx:
                    tempo_tracciato_val_orig = row_cells[col_tempo_tracciato_orig_idx - 1].value
                    tempo_tracciato_td_current_row = parse_time(tempo_tracciato_val_orig)


                current_output_col = 1
                for header_name_output in intestazioni_output:
                    cella_nuova = ws_nuovo.cell(row=nuova_riga, column=current_output_col)
                    
                    original_header_base_name = re.sub(r' \d+$', '', header_name_output) 
                    
                    original_col_idx_to_use = None

                    if original_header_base_name in intestazioni_fisse or original_header_base_name == "Tempo tracciato": # MODIFICA: Includi "Tempo tracciato"
                        if original_header_base_name in col_map_all_instances:
                            original_col_idx_to_use = col_map_all_instances[original_header_base_name][0]
                    
                    elif original_header_base_name in intestazioni_opzionali_entrata + intestazioni_opzionali_uscita:
                        match = re.search(r' (\d+)$', header_name_output)
                        suffix_num = int(match.group(1)) if match else 1 

                        count_current_type = 0
                        for col_type, col_idx_orig in all_entry_exit_cols:
                            if (col_type == "Entrata" and original_header_base_name == intestazioni_opzionali_entrata[0]) or \
                               (col_type == "Uscita" and original_header_base_name == intestazioni_opzionali_uscita[0]):
                                count_current_type += 1
                                if count_current_type == suffix_num:
                                    original_col_idx_to_use = col_idx_orig
                                    break
                    
                    if original_col_idx_to_use:
                        cella_orig = row_cells[original_col_idx_to_use - 1] 
                        cella_nuova.value = cella_orig.value

                        if cella_orig.has_style:
                            cella_nuova.font = copy(cella_orig.font)
                            cella_nuova.border = copy(cella_orig.border)
                            cella_nuova.fill = copy(cella_orig.fill)
                            cella_nuova.number_format = copy(cella_orig.number_format)
                            cella_nuova.protection = copy(cella_orig.protection)
                            cella_nuova.alignment = copy(cella_orig.alignment)
                        if cella_orig.hyperlink:
                            cella_nuova.hyperlink = copy(cella_orig.hyperlink)
                        if cella_orig.comment:
                            cella_nuova.comment = copy(cella_orig.comment)
                    
                    # LOGICA DI EVIDENZIAZIONE ORE LAVORATE
                    # Applica il colore rosso se "Ore lavorate" è diverso da "Tempo tracciato"
                    if header_name_output == "Ore lavorate":
                        if ore_lavorate_td_current_row != tempo_tracciato_td_current_row:
                            cella_nuova.fill = rosso_chiaro # Applica il riempimento rosso
                            # logger(f"Discrepanza per '{nome_foglio}' alla riga {nuova_riga}: Ore Lavorate '{format_timedelta(ore_lavorate_td_current_row)}' vs Tempo Tracciato '{format_timedelta(tempo_tracciato_td_current_row)}'")


                    # Logica per evidenziare la domenica (usa il valore nella nuova cella "Data")
                    if header_name_output == "Data" and isinstance(cella_nuova.value, str) and cella_nuova.value.strip().lower().startswith("dom"):
                        evidenzia_domenica = True # Usa la nuova variabile

                    current_output_col += 1

                if evidenzia_domenica: # Applica l'evidenziazione della domenica a tutta la riga
                    for c in range(1, current_output_col):
                        ws_nuovo.cell(row=nuova_riga, column=c).fill = rosso_chiaro
                    ultima_domenica = nuova_riga

                riga_map[nuova_riga] = evidenzia_domenica
                nuova_riga += 1
            
            # --- Calcolo delle somme ---
            col_somma = 1
            while any(ws_nuovo.cell(row=r, column=col_somma).value not in [None, ""] for r in range(1, nuova_riga)):
                col_somma += 1

            cella_somme_settimanali_header = ws_nuovo.cell(row=4, column=col_somma, value=intestazione_somme_settimanali)

            if "Ore lavorate" in col_idx_output_map:
                col_idx_riferimento = col_idx_output_map["Ore lavorate"]
                cella_riferimento_stile = ws_nuovo.cell(row=4, column=col_idx_riferimento) 

                if cella_riferimento_stile.has_style:
                    cella_somme_settimanali_header.font = copy(cella_riferimento_stile.font)
                    cella_somme_settimanali_header.border = copy(cella_riferimento_stile.border)
                    cella_somme_settimanali_header.fill = copy(cella_riferimento_stile.fill)
                    cella_somme_settimanali_header.number_format = copy(cella_riferimento_stile.number_format)
                    original_alignment_ref = copy(cella_riferimento_stile.alignment)
                    original_alignment_ref.wrap_text = True
                    cella_somme_settimanali_header.alignment = original_alignment_ref 
                    cella_somme_settimanali_header.protection = copy(cella_riferimento_stile.protection)
            else:
                cella_somme_settimanali_header.font = Font(bold=True) 
                cella_somme_settimanali_header.alignment = Alignment(wrap_text=True) 

            if "Ore lavorate" in intestazioni_output: 
                ore_col_idx = col_idx_output_map["Ore lavorate"]

                nome_dipendente_foglio = normalizza_nome(ws_nuovo.title)
                for nome_rif, ore_formattate in mappa_ore_settimanali.items(): 
                    if nome_rif in nome_dipendente_foglio or nome_dipendente_foglio in nome_rif:
                        col_per_ore_settimanali_contratto = col_somma + 1 
                        
                        cell_ore_sett_value = str(ore_formattate) 
                        
                        cell_ore_sett_contratto = ws_nuovo.cell(row=3, column=col_per_ore_settimanali_contratto) 
                        cell_ore_sett_contratto.value = "Ore" 
                        cell_ore_sett_contratto.font = Font(bold=False)
                        cell_ore_sett_contratto.alignment = Alignment(horizontal="left")
                        cell_ore_sett_contratto.alignment = Alignment(wrap_text=True)
                        cell_ore_sett_contratto.number_format = '@'

                        cell_ore_sett_val_cell = ws_nuovo.cell(row=4, column=col_per_ore_settimanali_contratto)
                        cell_ore_sett_val_cell.value = cell_ore_sett_value
                        cell_ore_sett_val_cell.font = Font(bold=False)
                        cell_ore_sett_val_cell.alignment = Alignment(horizontal="left")
                        cell_ore_sett_val_cell.alignment = Alignment(wrap_text=True)
                        cell_ore_sett_val_cell.number_format = '@' 

                        break
                else:
                    logger(f"⚠️ Nessuna corrispondenza trovata per '{ws_nuovo.title}' nel file ore settimanali.")
                
                for row_idx, is_dom in riga_map.items():
                    if is_dom:
                        start_r = max(5, row_idx - 6) 
                        tot = timedelta()
                        for r in range(start_r, row_idx + 1):
                            tot += parse_time(ws_nuovo.cell(row=r, column=ore_col_idx).value)
                        
                        cell_sum = ws_nuovo.cell(row=row_idx, column=col_somma)
                        cell_sum.value = format_timedelta(tot) 
                        cell_sum.font = Font(bold=True)
                        cell_sum.fill = rosso_chiaro
                        
                        cell_diff = ws_nuovo.cell(row=row_idx, column=col_somma + 1)
                        
                        diff_timedelta = tot - parse_time(cell_ore_sett_value)
                        
                        cell_diff.value = format_timedelta(diff_timedelta)
                        if diff_timedelta.total_seconds() < 0:
                            cell_diff.value = "-" + format_timedelta(abs(diff_timedelta))
                            cell_diff.fill = giallo_chiaro
                        else:
                            cell_diff.value = "+" + format_timedelta(diff_timedelta)
                            cell_diff.fill = verde_discrepanza

                for r in range(nuova_riga - 1, 1, -1):
                    v = ws_nuovo.cell(row=r, column=1).value
                    if isinstance(v, str) and "totale" in v.lower():
                        end_r = r - 1 
                        
                        start_r = ultima_domenica + 1 if ultima_domenica > 0 else 5 
                        
                        tot = timedelta()
                        for rr in range(start_r, end_r + 1):
                            tot += parse_time(ws_nuovo.cell(row=rr, column=ore_col_idx).value)
                        
                        cell_final = ws_nuovo.cell(row=end_r, column=col_somma)
                        cell_final.value = format_timedelta(tot)
                        cell_final.font = Font(bold=True)
                        cell_final.fill = rosso_chiaro
                        
                        cell_diff_final = ws_nuovo.cell(row=end_r, column=col_somma + 1)
                        final_diff_timedelta = tot - parse_time(cell_ore_sett_value)
                        if final_diff_timedelta.total_seconds() < 0:
                            cell_diff_final.value = "-" + format_timedelta(abs(final_diff_timedelta))
                            cell_diff_final.fill = giallo_chiaro
                        else:
                            cell_diff_final.value = "+" + format_timedelta(final_diff_timedelta)
                            cell_diff_final.fill = verde_discrepanza
                        break 
                
                info_totali_per_foglio[ws_nuovo.title] = (col_somma, r, ore_col_idx) 

        except Exception as e:
            logger(f"⚠️ Errore con il file {nome_file}: {e}")
            import traceback
            logger(traceback.format_exc()) 

        completati += 1
        if progress_callback:
            progress_callback(completati)
    
    total_ore_settimanali_complessive_di_tutti_i_fogli_rosse = timedelta(0) 

    for ws in wb_output.worksheets:
        info = info_totali_per_foglio.get(ws.title)
        
        if info:
            current_col_somma, week_total_cell_row, ore_col_idx_per_foglio = info
            
            somma_foglio, _ = trova_e_somma_celle_rosse_settimanali(ws, current_col_somma, rosso_chiaro)
            
            week_total_cell = ws.cell(row=week_total_cell_row, column=current_col_somma)
            week_total_cell.value = format_timedelta(somma_foglio)
            
            col_precedente = current_col_somma - 1
            if col_precedente >= 1:
                valore_col_precedente_str = ws.cell(row=week_total_cell_row, column=col_precedente).value
                valore_col_precedente_td = parse_time(valore_col_precedente_str)

                if somma_foglio == valore_col_precedente_td:
                    week_total_cell.fill = verde_discrepanza
                else:
                    week_total_cell.fill = rosso_scuro
                    logger(f"  ❌ DISCREPANZA in '{ws.title}': Totale ore elaborato ({format_timedelta(somma_foglio)}) ≠ Totale ore cartellino ({format_timedelta(valore_col_precedente_td)})")
            else:
                week_total_cell.fill = verde_discrepanza
            
            week_total_cell.font = Font(bold=True)
            
            total_ore_settimanali_complessive_di_tutti_i_fogli_rosse += somma_foglio

    num_fogli = len(wb_output.worksheets)
    
    # Easter Egg
    if num_fogli > 0: 
        num_easter_eggs = max(1, num_fogli // 10) 
        for foglio in random.sample(wb_output.worksheets, num_easter_eggs):
           foglio["E39"] = random.choice(["☀️", "🌵", "🌕", "🧸", "🌸"])

    for ws in wb_output.worksheets:
        ws.column_dimensions["A"].width = 15 
        
        for row in range(5, ws.max_row + 1): 
            ws.row_dimensions[row].height = 15
        
        info = info_totali_per_foglio.get(ws.title)
        if info:
            col_somma_attuale = info[0] 
            ws.column_dimensions[get_column_letter(col_somma_attuale)].width = 11 

        ws.delete_rows(1, 2) 

    wb_output.save(file_output)