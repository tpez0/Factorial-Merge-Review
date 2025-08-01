
import openpyxl
from openpyxl.styles import PatternFill, Font, Side, Border
from copy import copy

def confronta_file_cartellini(file_vecchio, file_nuovo, file_output, progress_callback=None):
    wb_old = openpyxl.load_workbook(file_vecchio, data_only=True)
    wb_new = openpyxl.load_workbook(file_nuovo)
    wb_output = openpyxl.Workbook()
    wb_output.remove(wb_output.active)

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    report_rows = []
    report_styles = []

    total_sheets = len(wb_old.sheetnames)
    for idx, sheet_name in enumerate(wb_old.sheetnames):
        if progress_callback:
            progress_callback(idx / total_sheets)
        if sheet_name not in wb_new.sheetnames:
            continue

        ws_old = wb_old[sheet_name]
        ws_new = wb_new[sheet_name]

        max_row = max(ws_old.max_row, ws_new.max_row)
        max_col = max(ws_old.max_column, ws_new.max_column)

        differenze_trovate = False
        ws_out = wb_output.create_sheet(title=sheet_name)

        # Copia larghezza colonne
        for col in ws_new.column_dimensions:
            ws_out.column_dimensions[col].width = ws_new.column_dimensions[col].width

        # Copia altezza righe
        for row in ws_new.row_dimensions:
            ws_out.row_dimensions[row].height = ws_new.row_dimensions[row].height

        for row in range(1, max_row + 1):
            riga_old = []
            riga_new = []
            stile_old = []
            stile_new = []
            modifiche_non_escluse = False

            for col in range(1, max_col + 1):
                if row == 37 and col == 5:
                    riga_old.append(None)
                    riga_new.append(None)
                    stile_old.append(None)
                    stile_new.append(None)
                    continue

                cell_old = ws_old.cell(row=row, column=col)
                cell_new = ws_new.cell(row=row, column=col)

                val_old = cell_old.value
                val_new = cell_new.value

                riga_old.append(val_old)
                riga_new.append(val_new)

                stile_old.append({
                    "font": copy(cell_old.font),
                    "fill": copy(cell_old.fill),
                    "border": copy(cell_old.border),
                    "number_format": copy(cell_old.number_format),
                    "alignment": copy(cell_old.alignment)
                })

                new_fill = red_fill if val_old != val_new else cell_new.fill
                stile_new.append({
                    "font": copy(cell_new.font),
                    "fill": copy(new_fill),
                    "border": copy(cell_new.border),
                    "number_format": copy(cell_new.number_format),
                    "alignment": copy(cell_new.alignment)
                })

                # Scrittura nel foglio confronto
                c_out = ws_out.cell(row=row, column=col, value=val_new)
                c_out.font = stile_new[-1]["font"]
                c_out.fill = stile_new[-1]["fill"]
                c_out.border = stile_new[-1]["border"]
                c_out.number_format = stile_new[-1]["number_format"]
                c_out.alignment = stile_new[-1]["alignment"]

                if val_old != val_new:
                    differenze_trovate = True
                    if not (row == 37 and col == 5):
                        modifiche_non_escluse = True

            if row >= 3 and riga_old != riga_new and modifiche_non_escluse:
                report_rows.append([sheet_name, "001"] + riga_old)
                report_rows.append([sheet_name, "002"] + riga_new)
                report_styles.append(["001"] + stile_old)
                report_styles.append(["002"] + stile_new)

        if not differenze_trovate:
            del wb_output[sheet_name]

    # Crea foglio report con stile (prima posizione)
    if report_rows:
        ws_report = wb_output.create_sheet("ReportDifferenze", 0)

        header = ["Foglio", "Versione"] + [
            wb_old.active.cell(row=2, column=col).value for col in range(1, wb_old.active.max_column + 1)
        ]
        ws_report.append(header)
        # Applica stile della riga 2 del primo foglio con modifiche all'intestazione
        reference_sheet = next((s for s in wb_output.sheetnames if s != "ReportDifferenze"), None)
        if reference_sheet:
            ws_ref = wb_output[reference_sheet]
            for col_idx, cell in enumerate(ws_report[1], start=1):
                ref_cell = ws_ref.cell(row=2, column=2)
                cell.font = copy(ref_cell.font)
                cell.fill = copy(ref_cell.fill)
                cell.border = copy(ref_cell.border)
                cell.number_format = copy(ref_cell.number_format)
                cell.alignment = copy(ref_cell.alignment)

        for i, row in enumerate(report_rows):
            ws_report.append(row)
            styles = report_styles[i][1:]  # salta "001"/"002"
            for col_idx, style in enumerate(styles, start=1):  # col 3 = dati, dopo "Foglio" (A) e "Versione" (B)
                c = ws_report.cell(row=ws_report.max_row, column=col_idx)
                c.font = style["font"]
                c.fill = style["fill"]
                c.border = style["border"]
                c.number_format = style["number_format"]
                c.alignment = style["alignment"]
                if row[1] == "002":
                    thin = Side(border_style="thin", color="000000")
                    for col_idx_border in range(1, len(row) + 1):
                        c = ws_report.cell(row=ws_report.max_row, column=col_idx_border)
                        existing = c.border
                        c.border = Border(
                            left=existing.left,
                            right=existing.right,
                            top=existing.top,
                            bottom=thin
                        )

        ws_report.row_dimensions[1].height = 24
        ws_report.column_dimensions["A"].width = 30
        ws_report.column_dimensions["B"].width = 11
        ws_report.column_dimensions["C"].width = 15
        for col in range(4, ws_report.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_report.column_dimensions[col_letter].width = 11

    wb_output.save(file_output)
