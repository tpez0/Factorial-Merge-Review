import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

def confronta_file_cartellini(file_vecchio, file_nuovo, file_output):
    wb_old = openpyxl.load_workbook(file_vecchio, data_only=True)
    wb_new = openpyxl.load_workbook(file_nuovo)
    wb_output = openpyxl.Workbook()
    wb_output.remove(wb_output.active)

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    report_differenze = []

    for sheet_name in wb_old.sheetnames:
        if sheet_name not in wb_new.sheetnames:
            continue

        ws_old = wb_old[sheet_name]
        ws_new = wb_new[sheet_name]

        max_row = max(ws_old.max_row, ws_new.max_row)
        max_col = max(ws_old.max_column, ws_new.max_column)

        differenze_trovate = False
        ws_out = wb_output.create_sheet(title=sheet_name)

        # Copia larghezza colonne
        for col in ws_new.column_dimensions:
            ws_out.column_dimensions[col].width = ws_new.column_dimensions[col].width

        # Copia altezza righe
        for row in ws_new.row_dimensions:
            ws_out.row_dimensions[row].height = ws_new.row_dimensions[row].height

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if row == 37 and col == 5:
                    continue  # ⬅️ Salta la cella E37
                cell_old = ws_old.cell(row=row, column=col).value
                cell_new = ws_new.cell(row=row, column=col).value
                c_new = ws_new.cell(row=row, column=col)
                c_out = ws_out.cell(row=row, column=col, value=cell_new)

                # Copia stile
                c_out.font = copy(c_new.font)
                c_out.fill = copy(c_new.fill)
                c_out.border = copy(c_new.border)
                c_out.number_format = copy(c_new.number_format)
                c_out.alignment = copy(c_new.alignment)

                if cell_old != cell_new:
                    differenze_trovate = True
                    c_out.fill = red_fill
                    data_val = ws_new.cell(row=row, column=1).value  # Valore della colonna A (colonna 1)
                    header = ws_old.cell(row=2, column=col).value if row != 2 else ""  # Intestazione della colonna
                    report_differenze.append([sheet_name, data_val, header, cell_old, cell_new])


        if not differenze_trovate:
            del wb_output[sheet_name]  # Elimina foglio se non ha modifiche

    if report_differenze:
        ws_report = wb_output.create_sheet("ReportDifferenze")
        ws_report.column_dimensions["A"].width = 30  # Imposta larghezza colonna A
        ws_report.column_dimensions["B"].width = 20
        ws_report.column_dimensions["C"].width = 20
        ws_report.column_dimensions["D"].width = 20

        ws_report.append(["Foglio", "Cella", "Colonna", "Vecchio valore", "Nuovo valore"])

        from openpyxl.styles import Font

        bold_font = Font(bold=True)

        for cell in ws_report[1]:  # Riga 1 (intestazione)
            cell.font = bold_font


        for riga in report_differenze:
            ws_report.append(riga)


    wb_output.save(file_output)
