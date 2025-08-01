import openpyxl
import os
from datetime import datetime, timedelta
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

def esegui_count_tot(file_input, file_output, progress_callback=None):
    wb = openpyxl.load_workbook(file_input)

    def parse_time(value):
        if isinstance(value, datetime):
            return value.time()
        elif isinstance(value, str):
            try:
                return datetime.strptime(value.strip(), "%H:%M").time()
            except:
                return None
        return None

    def td_to_hhmm(td):
        total_minutes = int(td.total_seconds() // 60)
        ore = total_minutes // 60
        minuti = total_minutes % 60
        return f"{ore}:{minuti:02d}"

    worksheets = wb.worksheets
    total_sheets = len(worksheets)
    riepilogo_data = []

    for sheet_index, ws in enumerate(worksheets, 1):
        max_col = ws.max_column
        target_col = max_col + 3
        col_letter = openpyxl.utils.get_column_letter(target_col)
        bold_font = Font(bold=True)

        # Scrive etichette
        ws[f"{col_letter}3"] = "Tot:"
        ws[f"{col_letter}4"] = "Notturni:"
        ws[f"{col_letter}5"] = "Domeniche:"
        ws[f"{col_letter}3"].font = bold_font
        ws[f"{col_letter}4"].font = bold_font
        ws[f"{col_letter}5"].font = bold_font

        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[2], 1) if cell.value}

        total_dur = timedelta()
        night_dur = timedelta()
        sunday_dur = timedelta()

        for row in ws.iter_rows(min_row=3):
            giorno = row[0].value
            is_sunday = isinstance(giorno, str) and giorno.lower().startswith("dom")

            for i in range(1, 4):
                ent_col = headers.get(f"Orario di entrata {i}")
                usc_col = headers.get(f"Orario di uscita {i}") or headers.get(f"Orario d'uscita {i}")
                if ent_col and usc_col:
                    start = parse_time(row[ent_col - 1].value)
                    end = parse_time(row[usc_col - 1].value)
                    if start and end and start != end:
                        dt_base = datetime(2000, 1, 1)
                        dt_start = datetime.combine(dt_base, start)
                        dt_end = datetime.combine(dt_base, end)
                        if dt_end <= dt_start:
                            dt_end += timedelta(days=1)
                        duration = dt_end - dt_start
                        total_dur += duration
                        if is_sunday:
                            sunday_dur += duration

                        current = dt_start
                        while current < dt_end:
                            next_min = min(current + timedelta(minutes=1), dt_end)
                            hour = current.time()
                            if hour >= datetime.strptime("22:00", "%H:%M").time() or hour < datetime.strptime("06:00", "%H:%M").time():
                                night_dur += next_min - current
                            current = next_min

        # Scrive risultati nel foglio
        for row_idx, duration in zip([3, 4, 5], [total_dur, night_dur, sunday_dur]):
            col_idx = target_col + 1
            while ws.cell(row=row_idx, column=col_idx).value:
                col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=td_to_hhmm(duration)).font = bold_font

        # Estrae info per il riepilogo
        nome = ws["B1"].value if ws["B1"].value else ws.title
        riepilogo_data.append([nome, td_to_hhmm(total_dur), td_to_hhmm(night_dur), td_to_hhmm(sunday_dur)])

        # Avanzamento barra
        if progress_callback:
            progress_callback(sheet_index / total_sheets)

        # Crea foglio Riepilogo come primo
        if "Riepilogo" in wb.sheetnames:
            del wb["Riepilogo"]
        ws_riepilogo = wb.create_sheet("Riepilogo", 0)

        # Intestazione
        intestazioni = ["Nome e Cognome", "Totale Ore", "Totale Notturni", "Totale Domeniche"]
        
        ws_riepilogo.append(intestazioni)
        bold_font = Font(bold=True)
        for cell in ws_riepilogo[1]:
            cell.font = bold_font
        
        for row in riepilogo_data:
            ws_riepilogo.append(row)

        # Larghezza colonne
        ws_riepilogo.column_dimensions["A"].width = 40
        for col in ["B", "C", "D"]:
            ws_riepilogo.column_dimensions[col].width = 17

        # Bordo inferiore a tutte le righe con dati
        thin_border = Border(bottom=Side(style="thin"))
        for row in ws_riepilogo.iter_rows(min_row=1, max_row=ws_riepilogo.max_row, max_col=4):
            for cell in row:
                cell.border = thin_border

    wb.save(file_output)
